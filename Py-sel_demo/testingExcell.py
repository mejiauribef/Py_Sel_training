import openpyxl

global book
book = openpyxl.load_workbook("C:/Users/f.mejia/Downloads/Product_Selectors_Regression_Testing.xlsx")
sheet = book.active
wb = book["Blood Pressure"]
test_dict = []
rows = wb.max_row
print(rows)
for i in range(rows-1):
    print(i)
    for j in range(1):
        test_selector_data = {}
        dict_key = wb.cell(row=1 , column= j+2).value
        dict_val = wb.cell(row=i+2 , column= j+2).value
        dict_key1 = wb.cell(row=1 , column= j+3).value
        dict_val1 = wb.cell(row=i+2 , column= j+3).value
        dict_key2 = wb.cell(row=1 , column= j+4).value
        dict_val2 = wb.cell(row=i+2 , column= j+4).value
        test_selector_data =  {dict_key:dict_val,dict_key1:dict_val1,dict_key2:dict_val2}
    test_dict.append(test_selector_data)
         

print(test_dict)